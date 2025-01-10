import cv2
import openpyxl
from pyzbar.pyzbar import decode
from datetime import datetime

# Load or create the Excel sheet
EXCEL_FILE = "attendance.xlsx"

def load_or_create_workbook():
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Date", "Time", "Name", "Barcode"])
        workbook.save(EXCEL_FILE)
    return workbook

# Check if a barcode has already been scanned
def is_barcode_scanned(sheet, barcode, date):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == date and row[3] == barcode:
            return True
    return False

# Add the user's entry to the Excel sheet
def add_entry(sheet, name, barcode):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    if not is_barcode_scanned(sheet, barcode, date):
        sheet.append([date, time, name, barcode])
        return True
    return False

# Main function to scan barcodes
def scan_barcode():
    workbook = load_or_create_workbook()
    sheet = workbook["Attendance"]

    # Open webcam
    cap = cv2.VideoCapture(0)
    print("Scanning for barcodes... Press 'q' to quit.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to access camera.")
            break

        for barcode in decode(frame):
            barcode_data = barcode.data.decode("utf-8")
            barcode_rect = barcode.rect

            # Simulating name lookup based on barcode
            name = f"User {barcode_data}"  # Replace with your lookup logic

            # Process attendance
            if add_entry(sheet, name, barcode_data):
                print(f"Attendance marked for {name} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                workbook.save(EXCEL_FILE)
                cv2.putText(frame, "Attendance Marked", (barcode_rect.left, barcode_rect.top - 10), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
            else:
                print(f"{name}'s attendance already marked for today.")
                cv2.putText(frame, "Already Marked", (barcode_rect.left, barcode_rect.top - 10), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

        # Display the camera feed
        cv2.imshow("Barcode Scanner", frame)

        # Quit on 'q' key press
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    scan_barcode()
